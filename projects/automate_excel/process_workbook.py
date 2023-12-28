import openpyxl as xl


def process_xl(file):
    wb = xl.load_workbook(file)  # loads the workbook
    sheet = wb['Sheet1']
    sheet.cell(1, 4).value = 'Corrected Price'  # adds a new column header

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # targets the cell in the 3rd column
        corrected_price = cell.value * 0.7  # runs the calculation
        # targets the cell in the 4th column for the new values
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price  # sets the new value to the cell

    return wb.save('file')  # saves the file
